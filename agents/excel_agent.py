"""
Excel Agent — Microsoft Excel automation via COM and openpyxl fallback.

Capabilities:
- Open/read/modify workbooks via pywin32 COM (preferred)
- Fall back to openpyxl when COM is unavailable
- Compute summaries, apply filters, create charts
- Export to CSV/PDF
"""
from __future__ import annotations

import os
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from agents.base_agent import BaseAgent
from models.schemas import AgentType, RiskLevel
from utils.logger import get_logger
from utils.helpers import ensure_dir, timestamped_filename, find_latest_file

logger = get_logger("agents.excel", "excel")

_COM_AVAILABLE = False
if sys.platform == "win32":
    try:
        import win32com.client as win32
        import pythoncom
        _COM_AVAILABLE = True
    except ImportError:
        logger.warning("pywin32 not available — using openpyxl fallback")

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    import pandas as pd
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl/pandas not available")


class ExcelAgent(BaseAgent):
    """Automates Microsoft Excel via COM with openpyxl fallback."""

    def __init__(self):
        self._xl_app = None       # COM Excel Application
        self._workbook = None     # COM Workbook
        self._wb_path: Optional[str] = None
        self._df_cache: Dict[str, Any] = {}  # sheet_name -> DataFrame
        super().__init__(AgentType.EXCEL)

    def _register_tools(self):
        self.register_tool("excel.open_workbook",   self.open_workbook,   "Open an Excel workbook", ["path"])
        self.register_tool("excel.close_workbook",  self.close_workbook,  "Close the active workbook")
        self.register_tool("excel.list_sheets",     self.list_sheets,     "List all sheet names")
        self.register_tool("excel.read_sheet",      self.read_sheet,      "Read all data from a sheet", ["sheet_name"])
        self.register_tool("excel.get_used_range",  self.get_used_range,  "Get the used range from a sheet", ["sheet_name"])
        self.register_tool("excel.read_range",      self.read_range,      "Read a specific cell range", ["sheet_name", "range_ref"])
        self.register_tool("excel.compute_summary", self.compute_summary, "Compute column statistics", ["sheet_name"])
        self.register_tool("excel.apply_filter",    self.apply_filter,    "Filter rows by condition", ["sheet_name", "column", "value"], risk_level=RiskLevel.MEDIUM)
        self.register_tool("excel.create_chart",    self.create_chart,    "Create a chart in Excel", ["sheet_name", "chart_type", "title"], risk_level=RiskLevel.MEDIUM)
        self.register_tool("excel.write_range",     self.write_range,     "Write data to a range", ["sheet_name", "start_cell", "data"], risk_level=RiskLevel.MEDIUM)
        self.register_tool("excel.apply_formatting",self.apply_formatting,"Apply cell formatting", ["sheet_name", "range_ref", "format_type"])
        self.register_tool("excel.save_workbook",   self.save_workbook,   "Save the workbook", risk_level=RiskLevel.MEDIUM)
        self.register_tool("excel.export_to_csv",   self.export_to_csv,   "Export a sheet to CSV", ["sheet_name", "output_path"])
        self.register_tool("excel.add_sheet",       self.add_sheet,       "Add a new sheet", ["sheet_name"])
        self.register_tool("excel.highlight_rows",  self.highlight_rows,  "Highlight rows matching a condition", ["sheet_name", "column", "condition", "threshold"])
        self.register_tool("excel.group_by",        self.group_by,        "Group rows by a column and aggregate another column", ["sheet_name", "group_column", "value_column"])

    # ─────────────────────────────────────────
    # Internal Helpers
    # ─────────────────────────────────────────

    @staticmethod
    def _sanitize_sheet_name(sheet_name: str) -> str:
        """
        Normalize a sheet name that may have been mangled by template resolution.

        The context manager sometimes resolves {{step_N.result.sheets}} to the
        string representation of a Python list, e.g. "['Sheet1']" or "['Sheet1', 'Sheet2']".
        This strips the list brackets and returns only the first sheet name.
        """
        s = str(sheet_name).strip()
        # Detect stringified list: starts with '[' and ends with ']'
        if s.startswith("[") and s.endswith("]"):
            # Parse safely — extract all quoted strings inside
            import re as _re
            names = _re.findall(r"'([^']*)'|\"([^\"]*)\"", s)
            flat = [a or b for a, b in names]
            if flat:
                logger.warning(
                    f"Sheet name looked like a list ({s!r}) — using first item: {flat[0]!r}"
                )
                return flat[0]
        return s

    @staticmethod
    def _resolve_column(df, col_name: str, prefer: str = "any") -> str:
        """
        Smart column resolution: exact → case-insensitive → substring → auto-detect.

        Parameters
        ----------
        df        : pandas DataFrame
        col_name  : The column name the planner guessed (may be vague like "category").
        prefer    : "numeric" — prefer a numeric column if auto-detecting.
                    "categorical" — prefer a string/object column if auto-detecting.
                    "any" — no preference.

        Returns the best-matching real column name, or raises ValueError.
        """
        import pandas as _pd

        available = list(df.columns)
        if not available:
            raise ValueError("DataFrame has no columns")

        # 1. Exact match
        if col_name in available:
            return col_name

        # 2. Case-insensitive exact match
        low = col_name.lower().replace(" ", "_")
        for c in available:
            if c.lower().replace(" ", "_") == low:
                logger.info(f"Column '{col_name}' matched case-insensitively to '{c}'")
                return c

        # 3. Substring / keyword match (e.g. "category" matches "product_category")
        for c in available:
            cl = c.lower().replace("_", " ")
            if low.replace("_", " ") in cl or cl in low.replace("_", " "):
                logger.info(f"Column '{col_name}' matched via substring to '{c}'")
                return c

        # 4. Auto-detect based on preference
        # Separate numeric and non-numeric columns
        numeric_cols = [c for c in available if _pd.api.types.is_numeric_dtype(df[c])]
        cat_cols = [c for c in available if not _pd.api.types.is_numeric_dtype(df[c])]

        if prefer == "numeric" and numeric_cols:
            # Pick the first numeric column (skip index-like columns)
            pick = next(
                (c for c in numeric_cols
                 if not any(skip in c.lower() for skip in ("id", "index", "row", "serial", "no", "s.no", "sno"))),
                numeric_cols[0],
            )
            logger.warning(f"Column '{col_name}' not found — auto-selected numeric column '{pick}'")
            return pick

        if prefer == "categorical" and cat_cols:
            # Pick the first non-numeric, non-date column
            pick = next(
                (c for c in cat_cols
                 if not any(skip in c.lower() for skip in ("date", "time", "timestamp", "id", "index", "row"))),
                cat_cols[0],
            )
            logger.warning(f"Column '{col_name}' not found — auto-selected categorical column '{pick}'")
            return pick

        # 5. Nothing matched — raise with helpful message
        raise ValueError(
            f"Column '{col_name}' not found and could not auto-detect a match.\n"
            f"Available columns: {available}"
        )

    # ─────────────────────────────────────────
    # Workbook Management
    # ─────────────────────────────────────────

    def open_workbook(self, path: str) -> Dict[str, Any]:
        """Open an Excel workbook. Tries COM first, then openpyxl."""
        resolved = str(Path(path).expanduser().resolve())

        # Reject Office lock files (e.g. ~$sales.xlsx created while Excel has the file open)
        if Path(resolved).name.startswith("~$"):
            raise ValueError(
                f"Cannot open Office lock file: {Path(resolved).name}\n"
                "This is a temporary file created by Excel/Word while the real file is open.\n"
                "Please close the file in Excel first, or use the real file (without the ~$ prefix)."
            )

        if not Path(resolved).exists():
            raise FileNotFoundError(f"Excel file not found: {resolved}")

        self._wb_path = resolved
        self._df_cache.clear()

        if _COM_AVAILABLE and not os.getenv("DEMO_MODE", "false").lower() == "true":
            return self._open_com(resolved)
        elif _OPENPYXL_AVAILABLE:
            return self._open_openpyxl(resolved)
        else:
            raise RuntimeError("No Excel library available (pywin32 or openpyxl)")

    def _open_com(self, path: str) -> Dict[str, Any]:
        """Open via COM — gives full Excel functionality."""
        try:
            pythoncom.CoInitialize()
            self._xl_app = win32.Dispatch("Excel.Application")
            self._xl_app.Visible = False
            self._xl_app.DisplayAlerts = False
            self._workbook = self._xl_app.Workbooks.Open(path)
            sheets = [self._workbook.Sheets(i+1).Name
                      for i in range(self._workbook.Sheets.Count)]
            logger.info(f"Opened via COM: {path} ({len(sheets)} sheets)")
            return {"path": path, "sheets": sheets, "method": "com", "status": "open"}
        except Exception as e:
            logger.warning(f"COM open failed ({e}), falling back to openpyxl")
            self._xl_app = None
            if _OPENPYXL_AVAILABLE:
                return self._open_openpyxl(path)
            raise

    def _open_openpyxl(self, path: str) -> Dict[str, Any]:
        """Open via openpyxl — read-only for .xlsx files."""
        self._workbook = openpyxl.load_workbook(path, data_only=True)
        sheets = self._workbook.sheetnames
        logger.info(f"Opened via openpyxl: {path} ({len(sheets)} sheets)")
        return {"path": path, "sheets": sheets, "method": "openpyxl", "status": "open"}

    def close_workbook(self, save: bool = False) -> Dict[str, Any]:
        """Close the active workbook."""
        if self._xl_app and self._workbook:
            try:
                if save and self._wb_path:
                    self._workbook.Save()
                self._workbook.Close(SaveChanges=False)
                self._xl_app.Quit()
            except Exception as e:
                logger.warning(f"COM close error: {e}")
            finally:
                self._xl_app = None
                self._workbook = None
        elif self._workbook and not self._xl_app:
            self._workbook.close()
            self._workbook = None
        return {"status": "closed"}

    # ─────────────────────────────────────────
    # Reading Data
    # ─────────────────────────────────────────

    def list_sheets(self) -> Dict[str, Any]:
        """Return all sheet names in the open workbook."""
        if not self._workbook:
            raise RuntimeError("No workbook open")
        if self._xl_app:
            sheets = [self._workbook.Sheets(i+1).Name
                      for i in range(self._workbook.Sheets.Count)]
        else:
            sheets = self._workbook.sheetnames
        return {"sheets": sheets, "count": len(sheets)}

    def get_used_range(self, sheet_name: str) -> Dict[str, Any]:
        """Get bounds of used data range in a sheet."""
        sheet_name = self._sanitize_sheet_name(sheet_name)
        if not self._workbook:
            raise RuntimeError("No workbook open")
        if self._xl_app:
            ws = self._workbook.Sheets(sheet_name)
            ur = ws.UsedRange
            return {
                "rows": ur.Rows.Count,
                "columns": ur.Columns.Count,
                "address": ur.Address,
                "sheet": sheet_name,
            }
        else:
            ws = self._workbook[sheet_name]
            return {
                "rows": ws.max_row,
                "columns": ws.max_column,
                "address": f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
                "sheet": sheet_name,
            }

    def read_sheet(self, sheet_name: str, max_rows: int = 1000) -> Dict[str, Any]:
        """Read all data from a sheet into a list-of-dicts format."""
        sheet_name = self._sanitize_sheet_name(sheet_name)
        if not self._workbook:
            raise RuntimeError("No workbook open")

        if self._xl_app:
            return self._read_sheet_com(sheet_name, max_rows)
        else:
            return self._read_sheet_openpyxl(sheet_name, max_rows)

    def _read_sheet_com(self, sheet_name: str, max_rows: int) -> Dict[str, Any]:
        ws = self._workbook.Sheets(sheet_name)
        ur = ws.UsedRange
        data = ur.Value  # Returns tuple of tuples

        if not data:
            return {"sheet": sheet_name, "headers": [], "rows": [], "row_count": 0}

        if isinstance(data[0], (str, int, float, type(None))):
            data = (data,)  # single row

        headers = [str(h) if h is not None else f"Col{i+1}"
                   for i, h in enumerate(data[0])]
        rows = []
        for row in data[1:min(len(data), max_rows + 1)]:
            rows.append({headers[i]: cell for i, cell in enumerate(row)})

        # Cache as DataFrame for summaries
        if _OPENPYXL_AVAILABLE:
            import pandas as pd
            self._df_cache[sheet_name] = self._clean_dataframe(pd.DataFrame(rows))

        return {
            "sheet": sheet_name,
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
        }

    def _read_sheet_openpyxl(self, sheet_name: str, max_rows: int) -> Dict[str, Any]:
        ws = self._workbook[sheet_name]
        rows_raw = list(ws.iter_rows(values_only=True, max_row=max_rows + 1))

        if not rows_raw:
            return {"sheet": sheet_name, "headers": [], "rows": [], "row_count": 0}

        headers = [str(h) if h is not None else f"Col{i+1}"
                   for i, h in enumerate(rows_raw[0])]
        rows = []
        for row in rows_raw[1:]:
            rows.append({headers[i]: val for i, val in enumerate(row)})

        if _OPENPYXL_AVAILABLE:
            import pandas as pd
            self._df_cache[sheet_name] = self._clean_dataframe(pd.DataFrame(rows))

        return {
            "sheet": sheet_name,
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
        }

    @staticmethod
    def _clean_dataframe(df) -> "pd.DataFrame":
        """
        Sanitise a freshly-loaded DataFrame:
        - Coerce columns that look numeric to proper numeric dtype
        - Strip whitespace from string columns
        - Drop fully-empty rows
        """
        import pandas as pd

        for col in df.columns:
            # Try to coerce object columns to numeric (handles "100", "200.5", None)
            if df[col].dtype == object:
                numeric = pd.to_numeric(df[col], errors="coerce")
                # Only convert if at least 50% of non-null values parsed as numbers
                non_null = df[col].dropna()
                numeric_non_null = numeric.dropna()
                if len(non_null) > 0 and len(numeric_non_null) / len(non_null) >= 0.5:
                    df[col] = numeric
                else:
                    # Clean string columns — strip whitespace, replace empty strings with NaN
                    df[col] = df[col].apply(
                        lambda x: x.strip() if isinstance(x, str) else x
                    )
                    df[col] = df[col].replace("", pd.NA)

        # Drop fully-empty rows
        df = df.dropna(how="all").reset_index(drop=True)
        return df

    @staticmethod
    def _safe_dict_rows(df) -> list:
        """
        Convert DataFrame rows to a list of dicts with NaN replaced by None.
        This ensures the result is JSON-serializable.
        """
        import math
        rows = df.to_dict("records")
        for row in rows:
            for k, v in row.items():
                if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                    row[k] = None
        return rows

    def read_range(self, sheet_name: str, range_ref: str) -> Dict[str, Any]:
        """Read a specific cell range."""
        if not self._workbook:
            raise RuntimeError("No workbook open")

        if self._xl_app:
            ws = self._workbook.Sheets(sheet_name)
            data = ws.Range(range_ref).Value
            if not isinstance(data, (list, tuple)):
                data = [[data]]
            return {"sheet": sheet_name, "range": range_ref, "data": [list(r) for r in data]}
        else:
            import openpyxl
            ws = self._workbook[sheet_name]
            data = [[cell.value for cell in row] for row in ws[range_ref]]
            return {"sheet": sheet_name, "range": range_ref, "data": data}

    # ─────────────────────────────────────────
    # Analysis
    # ─────────────────────────────────────────

    def compute_summary(
        self,
        sheet_name: str,
        numeric_only: bool = True,
        columns: List[str] = None,
        **kwargs,
    ) -> Dict[str, Any]:
        """Compute descriptive statistics for a sheet.

        Args:
            sheet_name: Name of the sheet to summarize.
            numeric_only: If True, only include numeric columns.
            columns: Optional list of columns to include. If omitted, all numeric columns are used.
            **kwargs: Extra LLM-generated kwargs are silently ignored.
        """
        sheet_name = self._sanitize_sheet_name(sheet_name)
        if sheet_name not in self._df_cache:
            self.read_sheet(sheet_name)

        if sheet_name not in self._df_cache:
            raise RuntimeError(f"Could not load data for sheet: {sheet_name}")

        df = self._df_cache[sheet_name]

        # Restrict to caller-specified columns if provided
        if columns:
            df = df[[c for c in columns if c in df.columns]]

        import math

        numeric_cols = df.select_dtypes(include="number").columns.tolist()

        summary = {}
        for col in numeric_cols:
            s = df[col].dropna()
            if s.empty:
                # All values are NaN/empty — skip or report zeros
                summary[col] = {
                    "total": 0, "mean": 0, "min": 0, "max": 0,
                    "count": 0, "std": 0, "empty": True,
                }
                continue

            # Safe float conversion — guard against NaN/Inf
            def _safe(v):
                f = float(v)
                return 0.0 if (math.isnan(f) or math.isinf(f)) else f

            summary[col] = {
                "total": _safe(s.sum()),
                "mean": _safe(s.mean()),
                "min": _safe(s.min()),
                "max": _safe(s.max()),
                "count": int(s.count()),
                "std": _safe(s.std()) if len(s) > 1 else 0.0,
            }

        top_rows = self._safe_dict_rows(df.head(5)) if not df.empty else []

        return {
            "sheet": sheet_name,
            "row_count": len(df),
            "columns": list(df.columns),
            "numeric_columns": numeric_cols,
            "summary": summary,
            "sample_rows": top_rows,
        }

    def group_by(
        self,
        sheet_name: str,
        group_column: str,
        value_column: str,
        agg: str = "sum",
        **kwargs,
    ) -> Dict[str, Any]:
        """Group rows by one column and aggregate another.

        Example: group_by(sheet_name='Sheet1', group_column='payment_mode',
                          value_column='total_amount', agg='sum')

        Args:
            sheet_name:   Sheet to read from.
            group_column: Column to group by (e.g. 'payment_mode').
            value_column: Column to aggregate (e.g. 'total_amount').
            agg:          Aggregation function — 'sum', 'mean', 'count', 'max', 'min'.
            **kwargs:     Extra LLM kwargs silently ignored.
        """
        sheet_name = self._sanitize_sheet_name(sheet_name)
        if sheet_name not in self._df_cache:
            self.read_sheet(sheet_name)

        df = self._df_cache.get(sheet_name)
        if df is None:
            raise RuntimeError(f"Sheet not loaded: {sheet_name}")

        group_column = self._resolve_column(df, group_column, prefer="categorical")
        value_column = self._resolve_column(df, value_column, prefer="numeric")

        agg_map = {
            "sum": "sum", "total": "sum",
            "mean": "mean", "average": "mean", "avg": "mean",
            "count": "count",
            "max": "max", "maximum": "max",
            "min": "min", "minimum": "min",
        }
        agg_fn = agg_map.get(agg.lower(), "sum")

        import math

        # Drop rows where the group column or value column is NaN — can't group by null
        clean = df.dropna(subset=[group_column, value_column]).copy()

        # Ensure value column is numeric
        import pandas as _pd
        clean[value_column] = _pd.to_numeric(clean[value_column], errors="coerce")
        clean = clean.dropna(subset=[value_column])

        if clean.empty:
            logger.warning(f"group_by: no valid data after dropping NaN rows")
            return {
                "group_column": group_column,
                "value_column": value_column,
                "aggregation": agg_fn,
                "groups": [],
                "group_count": 0,
                "grand_total": 0,
                "table_data": [],
                "table_headers": [group_column, f"{agg_fn}_{value_column}"],
            }

        grouped = clean.groupby(group_column)[value_column].agg(agg_fn).reset_index()
        grouped.columns = [group_column, f"{agg_fn}_{value_column}"]
        # Sort descending by value
        grouped = grouped.sort_values(f"{agg_fn}_{value_column}", ascending=False)

        rows = self._safe_dict_rows(grouped)
        grand_total_raw = float(grouped[f"{agg_fn}_{value_column}"].sum()) if agg_fn == "sum" else None
        total = 0.0 if (grand_total_raw is not None and math.isnan(grand_total_raw)) else grand_total_raw

        logger.info(
            f"group_by: {group_column} → {value_column} ({agg_fn}) — {len(rows)} groups"
        )
        return {
            "group_column": group_column,
            "value_column": value_column,
            "aggregation": agg_fn,
            "groups": rows,
            "group_count": len(rows),
            "grand_total": total,
            # Convenience — ready to pass directly to word.insert_table
            "table_data": rows,
            "table_headers": [group_column, f"{agg_fn}_{value_column}"],
        }

    def apply_filter(
        self,
        sheet_name: str,
        column: str,
        condition: str = "gt",
        value: Any = 0,
    ) -> Dict[str, Any]:
        """Filter rows matching a condition. Returns matching rows."""
        sheet_name = self._sanitize_sheet_name(sheet_name)
        if sheet_name not in self._df_cache:
            self.read_sheet(sheet_name)
        df = self._df_cache[sheet_name]

        ops = {"gt": ">", "gte": ">=", "lt": "<", "lte": "<=", "eq": "==", "ne": "!="}
        op = ops.get(condition, ">")

        try:
            filtered = df[df[column].apply(
                lambda x: eval(f"{x} {op} {repr(value)}")
                if x is not None and x == x else False
            )]
            rows = filtered.to_dict("records")
            return {
                "sheet": sheet_name,
                "column": column,
                "condition": f"{op} {value}",
                "matching_rows": len(rows),
                "rows": rows[:200],
            }
        except Exception as e:
            raise RuntimeError(f"Filter failed: {e}")

    # ─────────────────────────────────────────
    # Modification
    # ─────────────────────────────────────────

    def add_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """Add a new worksheet."""
        sheet_name = self._sanitize_sheet_name(sheet_name)
        if not self._workbook:
            raise RuntimeError("No workbook open")
        if self._xl_app:
            ws = self._workbook.Sheets.Add()
            ws.Name = sheet_name
        else:
            self._workbook.create_sheet(sheet_name)
        return {"added_sheet": sheet_name}

    def write_range(
        self,
        sheet_name: str,
        start_cell: str,
        data: List[List[Any]],
    ) -> Dict[str, Any]:
        """Write a 2D data array starting at start_cell."""
        sheet_name = self._sanitize_sheet_name(sheet_name)
        if not self._workbook:
            raise RuntimeError("No workbook open")
        if self._xl_app:
            ws = self._workbook.Sheets(sheet_name)
            rows = len(data)
            cols = max(len(r) for r in data) if data else 0
            rng = ws.Range(start_cell).Resize(rows, cols)
            rng.Value = data
        else:
            ws = self._workbook[sheet_name]
            for r_idx, row in enumerate(data):
                for c_idx, val in enumerate(row):
                    ws.cell(row=1 + r_idx, column=1 + c_idx, value=val)
        return {"written": True, "start": start_cell, "rows": len(data)}

    def apply_formatting(
        self,
        sheet_name: str,
        range_ref: str,
        format_type: str = "bold_header",
    ) -> Dict[str, Any]:
        """Apply basic formatting to a range."""
        sheet_name = self._sanitize_sheet_name(sheet_name)
        if not self._xl_app:
            return {"status": "skipped", "reason": "COM required for formatting"}
        ws = self._workbook.Sheets(sheet_name)
        rng = ws.Range(range_ref)

        if format_type == "bold_header":
            rng.Font.Bold = True
            rng.Interior.Color = 0x4472C4  # Blue
            rng.Font.Color = 0xFFFFFF      # White
        elif format_type == "highlight_yellow":
            rng.Interior.Color = 0xFFFF00
        elif format_type == "highlight_red":
            rng.Interior.Color = 0xFF0000
            rng.Font.Color = 0xFFFFFF
        elif format_type == "currency":
            rng.NumberFormat = "$#,##0.00"
        elif format_type == "date":
            rng.NumberFormat = "MM/DD/YYYY"

        return {"formatted": True, "range": range_ref, "type": format_type}

    def highlight_rows(
        self,
        sheet_name: str,
        column: str,
        condition: str = "gt",
        threshold: float = 0,
        color: str = "yellow",

    ) -> Dict[str, Any]:
        """Highlight entire rows where column meets condition."""
        sheet_name = self._sanitize_sheet_name(sheet_name)
        if not self._xl_app:
            return {"status": "skipped", "reason": "COM required for highlighting"}

        filtered = self.apply_filter(sheet_name, column, condition, threshold)
        ws = self._workbook.Sheets(sheet_name)
        colors = {"yellow": 0xFFFF00, "red": 0xFF0000, "green": 0x00FF00, "orange": 0xFFA500}
        fill_color = colors.get(color.lower(), 0xFFFF00)

        headers = [ws.Cells(1, c).Value for c in range(1, ws.UsedRange.Columns.Count + 1)]
        if column not in headers:
            return {"status": "error", "reason": f"Column {column} not found"}

        col_idx = headers.index(column) + 1
        highlighted = 0
        for row_idx in range(2, ws.UsedRange.Rows.Count + 1):
            cell_val = ws.Cells(row_idx, col_idx).Value
            try:
                if eval(f"{cell_val} {'>=' if condition == 'gt' else condition} {threshold}"):
                    ws.Rows(row_idx).Interior.Color = fill_color
                    highlighted += 1
            except Exception:
                pass

        return {"highlighted_rows": highlighted, "condition": f"{column} {condition} {threshold}"}

    def create_chart(
        self,
        sheet_name: str,
        chart_type: str = "bar",
        data_range: str = None,
        title: str = "Chart",
        output_image: str = None,
    ) -> Dict[str, Any]:
        """Create a chart in Excel (COM only) or export via matplotlib."""
        if self._xl_app and _COM_AVAILABLE:
            return self._create_chart_com(sheet_name, chart_type, data_range, title)
        elif _OPENPYXL_AVAILABLE:
            return self._create_chart_matplotlib(sheet_name, title, output_image)
        return {"status": "skipped", "reason": "No chart library available"}

    def _create_chart_com(self, sheet_name, chart_type, data_range, title) -> Dict:
        try:
            ws = self._workbook.Sheets(sheet_name)
            charts = ws.ChartObjects()
            chart_obj = charts.Add(100, 100, 400, 300)
            chart = chart_obj.Chart

            type_map = {"bar": -4100, "line": 4, "pie": 5, "column": 51}
            chart.ChartType = type_map.get(chart_type.lower(), -4100)

            if data_range:
                chart.SetSourceData(ws.Range(data_range))
            else:
                chart.SetSourceData(ws.UsedRange)

            chart.HasTitle = True
            chart.ChartTitle.Text = title
            return {"chart_created": True, "type": chart_type, "title": title}
        except Exception as e:
            return {"chart_created": False, "error": str(e)}

    def _create_chart_matplotlib(self, sheet_name, title, output_path) -> Dict:
        try:
            import matplotlib.pyplot as plt
            df = self._df_cache.get(sheet_name)
            if df is None:
                return {"status": "skipped", "reason": "No data cached"}

            numeric = df.select_dtypes(include="number")
            if numeric.empty:
                return {"status": "skipped", "reason": "No numeric data"}

            fig, ax = plt.subplots(figsize=(10, 6))
            numeric.head(20).plot(ax=ax, kind="bar")
            ax.set_title(title)
            ax.tick_params(axis='x', rotation=45)
            plt.tight_layout()

            if not output_path:
                output_path = str(Path.home() / "Desktop" / timestamped_filename("chart", "png"))
            plt.savefig(output_path, dpi=150, bbox_inches="tight")
            plt.close(fig)
            return {"chart_created": True, "path": output_path}
        except Exception as e:
            return {"chart_created": False, "error": str(e)}

    # ─────────────────────────────────────────
    # Save / Export
    # ─────────────────────────────────────────

    def save_workbook(self, path: str = None, format: str = "xlsx") -> Dict[str, Any]:
        """Save the active workbook."""
        if not self._workbook:
            raise RuntimeError("No workbook open")

        save_path = path or self._wb_path
        if not save_path:
            raise ValueError("No save path specified")

        if self._xl_app:
            fmt_map = {"xlsx": 51, "xls": 56, "csv": 6, "pdf": 57}
            fmt_code = fmt_map.get(format.lower(), 51)
            self._workbook.SaveAs(save_path, FileFormat=fmt_code)
        else:
            self._workbook.save(save_path)

        return {"saved": True, "path": save_path, "format": format}

    def export_to_csv(self, sheet_name: str, output_path: str) -> Dict[str, Any]:
        """Export a sheet to CSV using pandas."""
        sheet_name = self._sanitize_sheet_name(sheet_name)
        if sheet_name not in self._df_cache:
            self.read_sheet(sheet_name)
        df = self._df_cache.get(sheet_name)
        if df is None:
            raise RuntimeError(f"Sheet {sheet_name} not loaded")
        ensure_dir(Path(output_path).parent)
        df.to_csv(output_path, index=False)
        return {"exported": True, "path": output_path, "rows": len(df)}

    def __del__(self):
        """Cleanup COM objects on garbage collection."""
        try:
            if self._xl_app:
                self._workbook.Close(SaveChanges=False)
                self._xl_app.Quit()
        except Exception:
            pass
